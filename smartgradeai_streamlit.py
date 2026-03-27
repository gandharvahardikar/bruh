# -*- coding: utf-8 -*-
"""
SmartGrade AI — Streamlit Cloud Edition
Deploy: Push to GitHub → connect at share.streamlit.io
"""

import os
import re
import io
import cv2
import fitz  # PyMuPDF
import nltk
import shutil
import tempfile
import numpy as np
import pytesseract
import pandas as pd
import streamlit as st
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns

from PIL import Image
from pdf2image import convert_from_path
from difflib import SequenceMatcher
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from sentence_transformers import SentenceTransformer, util
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from doctr.io import DocumentFile
from doctr.models import ocr_predictor
from groq import Groq

# ─── Tesseract path ──────────────────────────────────────────────────────────
import platform
if platform.system() == "Windows":
    _t = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(_t):
        pytesseract.pytesseract.tesseract_cmd = _t

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="SmartGrade AI", page_icon="📝", layout="wide",
                   initial_sidebar_state="expanded")

# ─── CSS ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main-header {
    background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
    border-radius: 16px; padding: 2rem 2.5rem; margin-bottom: 1.5rem;
    box-shadow: 0 8px 32px rgba(0,0,0,0.25);
}
.main-header h1 { color: #fff; font-size: 2.4rem; font-weight: 800; margin: 0;
    text-shadow: 0 2px 8px rgba(0,0,0,0.3); }
.main-header p { color: rgba(255,255,255,0.8); font-size: 1.05rem; margin: .3rem 0 0; }
.metric-card {
    background: linear-gradient(145deg, #1a1a2e, #16213e);
    border-radius: 14px; padding: 1.4rem 1.2rem; text-align: center;
    box-shadow: 0 4px 18px rgba(0,0,0,0.18); border: 1px solid rgba(255,255,255,0.06);
}
.metric-card h3 { color: #a8dadc; font-size: .85rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: .04em; margin: 0 0 .5rem; }
.metric-card .value { color: #fff; font-size: 2rem; font-weight: 800; }
.metric-card .sub { color: rgba(255,255,255,0.5); font-size: .78rem; margin-top: .3rem; }
.grade-badge { display: inline-block; font-size: 2.6rem; font-weight: 800;
    padding: .5rem 1.6rem; border-radius: 14px; color: #fff;
    text-shadow: 0 2px 6px rgba(0,0,0,0.3); }
.grade-A { background: linear-gradient(135deg, #00b09b, #96c93d); }
.grade-B { background: linear-gradient(135deg, #36d1dc, #5b86e5); }
.grade-C { background: linear-gradient(135deg, #f7971e, #ffd200); color: #333; }
.grade-D { background: linear-gradient(135deg, #ed4264, #ffedbc); }
.grade-F { background: linear-gradient(135deg, #cb2d3e, #ef473a); }
.footer { text-align: center; padding: 1.2rem; color: #666;
    border-top: 1px solid rgba(255,255,255,0.08); margin-top: 2rem; font-size: .82rem; }
button[data-baseweb="tab"] { font-weight: 600 !important; font-size: .95rem !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>📝 SmartGrade AI</h1>
    <p>Intelligent answer-sheet grading — TrOCR · docTR · Tesseract · SBERT · Groq LLM</p>
</div>
""", unsafe_allow_html=True)


# ─── Model loading (cached) ─────────────────────────────────────────────────
@st.cache_resource(show_spinner="Loading NLP & OCR models (first run ≈ 2-3 min)…")
def load_models():
    nltk.download("punkt", quiet=True)
    nltk.download("punkt_tab", quiet=True)
    nltk.download("stopwords", quiet=True)
    return (
        set(stopwords.words("english")),
        ocr_predictor(pretrained=True),
        TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten"),
        VisionEncoderDecoderModel.from_pretrained("microsoft/trocr-base-handwritten"),
        SentenceTransformer("all-MiniLM-L6-v2"),
    )

STOP_WORDS, doctr_model, processor, trocr, sbert = load_models()


# ─── Core functions ──────────────────────────────────────────────────────────
def clean_image(path):
    img = cv2.imread(path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    denoised = cv2.fastNlMeansDenoising(enhanced, h=10)
    binarized = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, 11, 2)
    return Image.fromarray(binarized).convert("RGB")

def trocr_read(pil_img):
    px = processor(images=pil_img, return_tensors="pt").pixel_values
    ids = trocr.generate(px)
    return processor.batch_decode(ids, skip_special_tokens=True)[0]

def doctr_read(path):
    doc = DocumentFile.from_images(path)
    result = doctr_model(doc)
    return " ".join(w.value for p in result.pages for b in p.blocks
                    for l in b.lines for w in l.words)

def tesseract_read(pil_img):
    return pytesseract.image_to_string(pil_img)

def cosine_sim(t1, t2):
    if not t1.strip() or not t2.strip():
        return 0.0
    return float(util.cos_sim(sbert.encode(t1, convert_to_tensor=True),
                               sbert.encode(t2, convert_to_tensor=True)))

def best_ocr(path, reference):
    img = clean_image(path)
    candidates = {"TrOCR": trocr_read(img), "docTR": doctr_read(path),
                   "Tesseract": tesseract_read(img)}
    scored = {k: cosine_sim(v, reference) for k, v in candidates.items() if v.strip()}
    if not scored:
        return "", "None", 0.0
    winner = max(scored, key=scored.get)
    return candidates[winner], winner, scored[winner]

def pdf_to_text(path):
    doc = fitz.open(path)
    return "\n".join(page.get_text() for page in doc).strip()

def ocr_full_pdf(page_paths):
    full = ""
    for path in page_paths:
        doc = DocumentFile.from_images(path)
        result = doctr_model(doc)
        full += " ".join(w.value for p in result.pages for b in p.blocks
                          for l in b.lines for w in l.words) + "\n"
    return full

def parse_model_blocks(text):
    blocks = {}
    sub_q = re.compile(r'^\s*\(([A-Ea-e])\)\s+', re.MULTILINE)
    ans_p = re.compile(r'\bAns\b', re.IGNORECASE)
    mains = list(re.finditer(r'Q\.?\s*(\d+)', text, re.IGNORECASE))
    for mi, m in enumerate(mains):
        qn = m.group(1)
        qs, qe = m.end(), mains[mi+1].start() if mi+1 < len(mains) else len(text)
        qb = text[qs:qe]
        subs = list(sub_q.finditer(qb))
        for si, sm in enumerate(subs):
            key = f"Q{qn}{sm.group(1).upper()}"
            ss = sm.end()
            se = subs[si+1].start() if si+1 < len(subs) else len(qb)
            sb = qb[ss:se].strip()
            am = ans_p.search(sb)
            blocks[key] = sb[am.end():].strip() if am else sb
    return blocks

def semantic_split_student(ocr_text, model_blocks, min_chunk=60,
                            sim_threshold=0.30, min_total_chars=80, min_chunks_required=2):
    sentences = [s.strip() for s in re.split(r'\n{2,}|(?<=[.!?])\s+', ocr_text)
                 if len(s.strip()) > min_chunk]
    if not sentences:
        sentences = [s.strip() for s in ocr_text.split('\n') if len(s.strip()) > min_chunk]
    student_blocks = {k: "" for k in model_blocks}
    chunk_count = {k: 0 for k in model_blocks}
    for chunk in sentences:
        scores = {qk: cosine_sim(chunk, rt) for qk, rt in model_blocks.items()}
        best_key = max(scores, key=scores.get)
        if scores[best_key] >= sim_threshold:
            student_blocks[best_key] += " " + chunk
            chunk_count[best_key] += 1
    return {k: v.strip() for k, v in student_blocks.items()
            if len(v.strip()) >= min_total_chars and chunk_count[k] >= min_chunks_required}

def keyword_score(student, reference):
    def kw(t):
        return set(w for w in word_tokenize(t.lower()) if w.isalpha() and w not in STOP_WORDS)
    rk, sk = kw(reference), kw(student)
    return len(sk & rk) / len(rk) if rk else 0.0

def compute_score(student_text, ref_text, max_marks=5):
    s_sem = cosine_sim(student_text, ref_text)
    s_kw = keyword_score(student_text, ref_text)
    s_seq = SequenceMatcher(None, student_text, ref_text).ratio()
    w = 0.6*s_sem + 0.2*s_kw + 0.2*s_seq
    return {"semantic": round(s_sem, 4), "keyword": round(s_kw, 4),
            "sequence": round(s_seq, 4), "weighted": round(w, 4),
            "final": round(w * max_marks, 2), "max": max_marks, "flag": w < 0.5}

def llm_evaluate(student_text, ref_text, qkey, max_marks, groq_client, groq_model):
    prompt = f"""You are a chill university professor grading a Computer Networks exam.

Question ID: {qkey}
Maximum Marks: {max_marks}

Model Answer (reference):
{ref_text}

Student Answer:
{student_text}

Grading Rules:
- Award marks based on conceptual correctness, not exact wording
- If the student explains the same concept differently or uses simpler language, award full/near-full marks
- Partial credit for partially correct answers
- Deduct marks only for factually wrong statements or missing key concepts
- The student answer may have OCR noise — ignore spelling errors, focus on meaning
- Do NOT give 0 unless the answer is completely blank or entirely wrong topic

Respond in this exact format only, no extra text:
SCORE: <number out of {max_marks}, can be decimal like 3.5>
RELEVANT: Yes or No
MISSING: <key concepts missing, or None>
REASON: <one sentence justification>"""
    try:
        resp = groq_client.chat.completions.create(
            model=groq_model, messages=[{"role": "user", "content": prompt}],
            temperature=0.1, max_tokens=150)
        raw = resp.choices[0].message.content.strip()
        score = min(float(re.findall(r"[\d.]+", raw.split("SCORE:")[1].split("\n")[0])[0]), max_marks)
        return {"llm_score": round(score, 2),
                "relevant": "yes" in raw.split("RELEVANT:")[1].split("\n")[0].lower(),
                "missing": raw.split("MISSING:")[1].split("\n")[0].strip(),
                "reason": raw.split("REASON:")[1].strip().split("\n")[0], "raw": raw}
    except Exception as e:
        return {"llm_score": None, "relevant": True, "missing": "parse error",
                "reason": str(e), "raw": ""}


# ─── Chart generation ────────────────────────────────────────────────────────
def generate_charts(records):
    sns.set_theme(style="darkgrid", palette="muted")
    fig = plt.figure(figsize=(20, 18))
    fig.suptitle("SmartGrade AI — Analytics Dashboard", fontsize=18, fontweight="bold", y=0.98)
    qs = [r["question"] for r in records]
    scores = [r["auto_score"] for r in records]
    sem = [r["semantic_sim"] for r in records]
    kw = [r["keyword_cov"] for r in records]
    seq = [r["sequence_sim"] for r in records]
    clrs = ["#2ecc71" if not r["needs_review"] else "#e74c3c" for r in records]

    # 1 Score per Q
    ax = fig.add_subplot(3,3,1)
    bars = ax.bar(qs, scores, color=clrs, edgecolor="white")
    ax.set_ylim(0, max(scores)+1 if scores else 6); ax.set_title("Score per Question", fontweight="bold")
    for b, s in zip(bars, scores):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+.05, f"{s:.1f}", ha="center", fontsize=9)
    ax.legend(handles=[mpatches.Patch(color="#2ecc71", label="OK"),
                        mpatches.Patch(color="#e74c3c", label="Review")], fontsize=8)

    # 2 Semantic
    ax = fig.add_subplot(3,3,2)
    ax.bar(qs, sem, color="#3498db"); ax.axhline(.7, color="orange", ls="--", label="0.7")
    ax.set_ylim(0,1.05); ax.set_title("Semantic Similarity", fontweight="bold"); ax.legend(fontsize=8)

    # 3 Keyword
    ax = fig.add_subplot(3,3,3)
    ax.bar(qs, kw, color="#9b59b6"); ax.axhline(.5, color="orange", ls="--", label="0.5")
    ax.set_ylim(0,1.05); ax.set_title("Keyword Coverage", fontweight="bold"); ax.legend(fontsize=8)

    # 4 Sequence
    ax = fig.add_subplot(3,3,4)
    ax.bar(qs, seq, color="#1abc9c"); ax.set_ylim(0,1.05)
    ax.set_title("Sequence Similarity", fontweight="bold")

    # 5 Grouped
    ax = fig.add_subplot(3,3,5)
    x = np.arange(len(qs)); w = .25
    ax.bar(x-w, sem, w, label="Semantic", color="#3498db")
    ax.bar(x, kw, w, label="Keyword", color="#9b59b6")
    ax.bar(x+w, seq, w, label="Sequence", color="#1abc9c")
    ax.set_xticks(x); ax.set_xticklabels(qs, fontsize=8, rotation=45)
    ax.set_title("All Metrics Grouped", fontweight="bold"); ax.legend(fontsize=8)

    # 6 Relevance pie
    ax = fig.add_subplot(3,3,6)
    rel = sum(1 for r in records if r["relevant"]); nrel = len(records)-rel
    if rel+nrel > 0:
        ax.pie([rel, nrel], labels=["Relevant","Not Relevant"], colors=["#2ecc71","#e74c3c"],
               autopct="%1.0f%%", startangle=90)
    ax.set_title("Answer Relevance", fontweight="bold")

    # 7 Histogram
    ax = fig.add_subplot(3,3,7)
    ax.hist(scores, bins=10, color="#e67e22", edgecolor="white")
    ax.axvline(np.mean(scores), color="red", ls="--", label=f"Mean {np.mean(scores):.1f}")
    ax.set_title("Score Distribution", fontweight="bold"); ax.legend(fontsize=8)

    # 8 Radar
    ax = fig.add_subplot(3,3,8, polar=True)
    cats = ["Semantic","Keyword","Sequence","Avg Score"]
    vals = [np.mean(sem), np.mean(kw), np.mean(seq),
            np.mean(scores)/max(r["out_of"] for r in records) if records else 0]
    vals += vals[:1]
    ang = np.linspace(0, 2*np.pi, 4, endpoint=False).tolist() + [0]
    ax.plot(ang, vals, "o-", color="#3498db"); ax.fill(ang, vals, alpha=.25, color="#3498db")
    ax.set_xticks(ang[:-1]); ax.set_xticklabels(cats, fontsize=8)
    ax.set_title("Radar Chart", fontweight="bold", pad=15)

    # 9 Summary table
    ax = fig.add_subplot(3,3,9); ax.axis("off")
    tot = sum(scores); mx = sum(r["out_of"] for r in records)
    pct = round(tot/mx*100,1) if mx else 0
    rows = [["Total Score", f"{tot:.1f}/{mx}"], ["Percentage", f"{pct}%"],
            ["Questions", str(len(records))], ["Avg Score", f"{np.mean(scores):.2f}"],
            ["Avg Semantic", f"{np.mean(sem):.2f}"], ["Avg Keyword", f"{np.mean(kw):.2f}"],
            ["Needs Review", str(sum(1 for r in records if r["needs_review"]))]]
    tbl = ax.table(cellText=rows, colLabels=["Metric","Value"], cellLoc="center",
                    loc="center"); tbl.auto_set_font_size(False); tbl.set_fontsize(11)
    tbl.scale(1.2, 1.8)
    for (row, col), cell in tbl.get_celld().items():
        if row == 0:
            cell.set_facecolor("#2c3e50"); cell.set_text_props(color="white", fontweight="bold")
        elif row % 2 == 0:
            cell.set_facecolor("#f2f3f4")
    ax.set_title("Summary Statistics", fontweight="bold")

    plt.tight_layout(rect=[0,0,1,.97])
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight"); plt.close(); buf.seek(0)
    return buf


# ─── Grading pipeline ────────────────────────────────────────────────────────
def run_grading(hw_bytes, mk_bytes, qp_bytes, student_name, groq_api_key,
                max_marks, groq_model_name):
    work = tempfile.mkdtemp(prefix="smartgrade_")
    hw_p, mk_p, qp_p = [os.path.join(work, n) for n in
                          ("handwritten.pdf","model_key.pdf","questions.pdf")]
    pages_dir = os.path.join(work, "pages"); os.makedirs(pages_dir)
    for p, b in [(hw_p, hw_bytes), (mk_p, mk_bytes), (qp_p, qp_bytes)]:
        with open(p, "wb") as f: f.write(b)

    bar = st.progress(0, text="Converting PDF → images…")
    raw_pages = convert_from_path(hw_p, dpi=300)
    page_paths = []
    for i, pg in enumerate(raw_pages):
        out = os.path.join(pages_dir, f"page_{i+1}.png"); pg.save(out, "PNG"); page_paths.append(out)

    bar.progress(10, text=f"Running OCR on {len(page_paths)} pages…")
    model_raw = pdf_to_text(mk_p)
    student_raw = ocr_full_pdf(page_paths)

    bar.progress(40, text="Parsing & matching answers…")
    model_blocks = parse_model_blocks(model_raw)
    student_blocks = semantic_split_student(student_raw, model_blocks)

    gc = Groq(api_key=groq_api_key) if groq_api_key and groq_api_key.strip() else None

    records = []; total_q = max(len(model_blocks), 1)
    for i, (qk, ref) in enumerate(model_blocks.items()):
        bar.progress(45 + int(45*i/total_q), text=f"Grading {qk}…")
        stxt = student_blocks.get(qk, "")
        if not stxt: continue
        llm_r = (llm_evaluate(stxt, ref, qk, max_marks, gc, groq_model_name) if gc
                 else {"llm_score": None, "relevant": True, "missing": "N/A",
                       "reason": "LLM skipped (no key)", "raw": ""})
        sb_r = compute_score(stxt, ref, max_marks)
        ls = llm_r["llm_score"] or 0.0; ss = sb_r["final"]; fs = max(ls, ss)
        records.append({"question": qk, "auto_score": fs, "llm_score": ls,
            "sbert_score": ss, "winner": "LLM" if ls >= ss else "SBERT",
            "out_of": max_marks, "relevant": llm_r["relevant"],
            "missing": llm_r["missing"], "reason": llm_r["reason"],
            "needs_review": fs < 2.0, "semantic_sim": sb_r["semantic"],
            "keyword_cov": sb_r["keyword"], "sequence_sim": sb_r["sequence"]})

    bar.progress(93, text="Generating charts…")
    chart = generate_charts(records) if records else None
    df = pd.DataFrame(records)
    bar.progress(100, text="✅ Done!")
    shutil.rmtree(work, ignore_errors=True)
    return records, chart, df


# ─── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Upload PDFs")
    hw_file = st.file_uploader("1️⃣  Handwritten Answer Sheet", type=["pdf"], key="hw")
    mk_file = st.file_uploader("2️⃣  Model Answer Key", type=["pdf"], key="mk")
    qp_file = st.file_uploader("3️⃣  Question Paper", type=["pdf"], key="qp")
    st.markdown("---")
    st.markdown("### ⚙️ Settings")
    student_name = st.text_input("Student Name *", placeholder="e.g. Rahul Patil")
    # API key: use secrets if deployed, otherwise manual input
    default_key = st.secrets.get("GROQ_API_KEY", "") if hasattr(st, "secrets") else ""
    groq_key = st.text_input("Groq API Key", type="password", value=default_key,
                              placeholder="gsk_…",
                              help="Get free key → [console.groq.com](https://console.groq.com)")
    groq_model = st.selectbox("Groq Model", ["llama-3.1-8b-instant",
        "llama-3.3-70b-versatile", "mixtral-8x7b-32768", "gemma2-9b-it"])
    max_marks = st.slider("Max Marks per Question", 1, 10, 5, 1)
    st.markdown("---")
    run_btn = st.button("⚡ Run AI Grading", type="primary", use_container_width=True)


# ─── Execute ─────────────────────────────────────────────────────────────────
if run_btn:
    if not hw_file or not mk_file or not qp_file:
        st.error("❌ Upload all 3 PDFs first."); st.stop()
    if not student_name.strip():
        st.error("❌ Enter student name."); st.stop()
    records, chart_buf, df = run_grading(
        hw_file.read(), mk_file.read(), qp_file.read(),
        student_name.strip(), groq_key, max_marks, groq_model)
    st.session_state.update(records=records, chart_buf=chart_buf, df=df,
                             student_name=student_name.strip(), max_marks=max_marks)


# ─── Display results ─────────────────────────────────────────────────────────
if "records" in st.session_state and st.session_state["records"]:
    records = st.session_state["records"]
    chart_buf = st.session_state["chart_buf"]
    df = st.session_state["df"]
    sname = st.session_state["student_name"]

    total_score = round(df["auto_score"].sum(), 2)
    total_max = int(df["out_of"].sum())
    pct = round(total_score/total_max*100, 1) if total_max else 0
    grade = "A" if pct>=90 else "B" if pct>=75 else "C" if pct>=60 else "D" if pct>=40 else "F"

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["📊 Results", "📈 Analytics", "🗂 Data Table", "💾 Downloads", "❓ Help"])

    with tab1:
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.markdown(f'<div class="metric-card"><h3>Student</h3>'
                    f'<div class="value" style="font-size:1.3rem">{sname}</div></div>',
                    unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card"><h3>Total Score</h3>'
                    f'<div class="value">{total_score}</div>'
                    f'<div class="sub">out of {total_max}</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="metric-card"><h3>Percentage</h3>'
                    f'<div class="value">{pct}%</div></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="metric-card"><h3>Grade</h3>'
                    f'<div class="grade-badge grade-{grade}">{grade}</div></div>',
                    unsafe_allow_html=True)
        flagged = int(df["needs_review"].sum())
        c5.markdown(f'<div class="metric-card"><h3>Needs Review</h3>'
                    f'<div class="value" style="color:{"#e74c3c" if flagged else "#2ecc71"}">'
                    f'{flagged}</div><div class="sub">of {len(records)} questions</div></div>',
                    unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("Per-Question Breakdown")
        for r in records:
            status = "⚠️ Review" if r["needs_review"] else "✅ OK"
            cols = st.columns([2,1,1,1,1,1,3])
            cols[0].markdown(f"**{r['question']}**")
            cols[1].write(f"{r['auto_score']}/{r['out_of']}")
            cols[2].write(f"🏆 {r['winner']}")
            cols[3].write(f"Sem: {r['semantic_sim']:.2f}")
            cols[4].write(f"Kw: {r['keyword_cov']:.2f}")
            cols[5].write(status)
            cols[6].caption(r.get("reason","")[:100])

    with tab2:
        if chart_buf:
            st.image(chart_buf, caption="9-Panel Analytics Dashboard", use_container_width=True)
        else:
            st.info("No data to chart.")

    with tab3:
        disp = df[["question","auto_score","llm_score","sbert_score","winner","out_of",
                    "semantic_sim","keyword_cov","sequence_sim","relevant","missing",
                    "reason","needs_review"]].copy()
        disp.columns = ["Question","Final","LLM","SBERT","Grader","Max","Semantic",
                         "Keyword","Sequence","Relevant","Missing","Reasoning","Review"]
        st.dataframe(disp, use_container_width=True, hide_index=True)

    with tab4:
        st.markdown("### 📥 Download Results")
        ca, cb = st.columns(2)
        ca.download_button("📄 CSV", df.to_csv(index=False).encode(), "smartgrade_results.csv",
                            "text/csv", use_container_width=True)
        if chart_buf:
            chart_buf.seek(0)
            cb.download_button("📊 Chart PNG", chart_buf, "smartgrade_analytics.png",
                                "image/png", use_container_width=True)
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            xbuf = io.BytesIO()
            with pd.ExcelWriter(xbuf, engine="openpyxl") as w:
                df.to_excel(w, index=False, sheet_name="Report")
                ws = w.sheets["Report"]
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="2C3E50")
                    cell.alignment = Alignment(horizontal="center")
            xbuf.seek(0)
            st.download_button("📗 Excel", xbuf, "smartgrade_report.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except ImportError:
            st.caption("Install openpyxl for Excel export.")

    with tab5:
        st.markdown("""
## How to Use
1. Upload **3 PDFs** in the sidebar
2. Enter **student name** + optional **Groq API key**
3. Click **⚡ Run AI Grading** → wait 2-5 min
4. Check tabs for results, analytics, downloads

---
| Method | How |
|--------|-----|
| **SBERT** | Semantic similarity + keyword + sequence match |
| **LLM** | Groq evaluates relevance + awards marks |
| **Final** | `max(LLM, SBERT)` — best of both |

✅ OK = score ≥ 2.0 &nbsp;&nbsp; ⚠️ Review = score < 2.0

Free Groq key → [console.groq.com](https://console.groq.com)
        """)

else:
    st.markdown("""
    <div style="text-align:center; padding:4rem 2rem;">
        <h2 style="color:#a8dadc;">👋 Welcome to SmartGrade AI</h2>
        <p style="color:#888; font-size:1.1rem; max-width:600px; margin:auto;">
            Upload your PDFs in the sidebar and click <strong>⚡ Run AI Grading</strong> to begin.
        </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div class="footer">SmartGrade AI — TrOCR · docTR · Tesseract · SBERT · Groq LLM</div>',
            unsafe_allow_html=True)
